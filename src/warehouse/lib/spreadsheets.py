import logging
from pathlib import Path

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation

# Get logging process
log = logging.getLogger(Path(__file__).stem)


def apply_worksheet_validation_rule(worksheet, validation_dict: dict) -> None:
    """
    Restores spreadsheet data validation rules from a dictionary

    worksheet: openpyxl workbook/sheet
    validation_dict: dictionary containing data validation rules

    """
    # Define validation refs
    ref_type = validation_dict.get("ref_type")
    ref_range = validation_dict.get("ref_range")

    # Determine which cell or cells it should be applied to
    cell_range = validation_dict.get("apply_to_cell")

    log.debug(
        f"         Applying {ref_type} with lookup range {ref_range} to range {cell_range}"
    )

    # Range of cells
    if ":" in cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                each_cell = worksheet.cell(row=row, column=col)
                data_validation = DataValidation(type=ref_type, formula1=ref_range)
                worksheet.add_data_validation(data_validation)
                data_validation.add(each_cell)
    # Single cell
    else:
        data_validation = DataValidation(type=ref_type, formula1=ref_range)
        worksheet.add_data_validation(data_validation)
        data_validation.add(worksheet[cell_range])


def apply_worksheet_conditional_formatting(worksheet, formatting_dict: dict) -> None:
    """
    Restores spreadsheet data validation rules from a dictionary

    worksheet: openpyxl workbook/sheet
    validation_dict: dictionary containing data validation rules

    """

    # Define formatting references
    value_to_compare = formatting_dict.get("value_to_compare")
    operator = formatting_dict.get("operator")
    colour = formatting_dict.get("colour")

    # Build the format
    fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")

    # Determine which cell or cells it should be applied to
    cell_range = formatting_dict.get("apply_to_cell")

    log.debug(
        f"         Applying {colour} with {value_to_compare} to range {cell_range}"
    )

    # Range of cells
    if ":" in cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                each_cell = worksheet.cell(row=row, column=col).coordinate
                worksheet.conditional_formatting.add(
                    each_cell,
                    CellIsRule(
                        operator=operator,
                        formula=[str(value_to_compare)],
                        stopIfTrue=True,
                        fill=fill,
                    ),
                )
    # Single cell
    else:
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator=operator,
                formula=[str(value_to_compare)],
                stopIfTrue=True,
                fill=fill,
            ),
        )


def extract_values_from_named_range(workbook, range_name: str) -> list:
    """
    Extracts values from a named range in an Excel workbook.

    workbook: openpyxl workbook object
    range_name: name of the range to extract values from

    Returns a list of values from the named range.
    """
    named_range = workbook.defined_names[range_name]
    # workbook.defined_names["exp_version"]

    if not named_range:
        log.error(f"Named range '{range_name}' not found in the workbook.")
        return []

    values = []
    for sheetname, cell_address in named_range.destinations:
        sheet = workbook[sheetname]
        cell = sheet[cell_address]
        values.append(cell.value)

    return values
