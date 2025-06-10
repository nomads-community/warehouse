import logging
from pathlib import Path

import click
import yaml
from openpyxl import load_workbook

from warehouse.lib.exceptions import DataFormatError
from warehouse.lib.general import identify_files_by_search, pad_list, produce_dir
from warehouse.lib.logging import divider, identify_cli_command
from warehouse.lib.regex import Regex_patterns
from warehouse.lib.spreadsheets import (
    apply_worksheet_conditional_formatting,
    apply_worksheet_validation_rule,
)

# Resolve file / folder locations irrespective of cwd
script_dir = Path(__file__).parent.resolve()
templates_dir = script_dir.parent.parent.parent / "templates"
# Identify and load targets dict from YAML file (assuming the file exists)
group_details_yaml = script_dir / "group_details.yml"
data_validations_yaml = script_dir / "data_validations.yml"
conditional_formatting_yaml = script_dir / "conditional_formatting.yml"


@click.command(
    short_help="Update NOMADS template files with group specific usernames and projects"
)
@click.option(
    "-o",
    "--output_folder",
    type=Path,
    required=False,
    help="Path to folder where the updated templates should be output to",
)
@click.option(
    "-g",
    "--group_name",
    type=str,
    required=False,
    help="Name of group to use",
)
@click.option(
    "-l",
    "--list_groups",
    is_flag=True,
    help="List all groups available",
)
def templates(group_name: str, output_folder: Path, list_groups: bool):
    """
    Update NOMADS templates with user and project names
    """

    # Set up child log
    log = logging.getLogger("templates_commands")
    log.info(divider)
    log.debug(identify_cli_command())

    # Load group details from YAML file
    with open(group_details_yaml, "r") as f:
        details = yaml.safe_load(f)

    # Load data validation details from YAML file
    with open(data_validations_yaml, "r") as f:
        validations = yaml.safe_load(f)

    # Load conditional formatting details from YAML file
    with open(conditional_formatting_yaml, "r") as f:
        formatting_rules = yaml.safe_load(f)

    # List group options
    if list_groups:
        groups = ", ".join(list(details.keys()))
        log.info(f"Available groups are: {groups}")
        log.info(divider)
        return

    # Extract group names
    grp_details = details.get(group_name)
    if grp_details is None:
        raise DataFormatError(
            f"-g '{group_name}' not found. Options are {list(details.keys())}."
        )

    if output_folder is None:
        log.info("You have not defined an output folder (-o) option")
        log.info(divider)
        return

    # Identify all template files
    template_fns = identify_files_by_search(templates_dir, Regex_patterns.EXCEL_FILE)

    # Create the output folder
    produce_dir(output_folder)

    # For each template change the names, initials and projects
    for template_fn in template_fns:
        # Load the workbook
        workbook = load_workbook(template_fn)
        # Select the correct worksheet
        worksheet = workbook["reference"]

        for dict_key, col_num in zip(grp_details.keys(), [9, 10, 12]):
            # Extract the correct values to enter
            details = pad_list(grp_details, dict_key, 6)
            # Names are in I3-I8, Initials in J3 to J8 and Projects in L3 to L8
            for count, xl_row in enumerate(range(3, 8)):
                # Define the cell to edit
                cell = worksheet.cell(row=xl_row, column=col_num)
                # Assign the new value to the cell
                cell.value = details[count]

        # Then restore the data validation logic that is somehow overwritten when replacing the above details
        log.debug(f"{template_fn.stem}")

        for sheetname, validation in validations.get(template_fn.stem, {}).items():
            # Load worksheet
            worksheet = workbook[sheetname]
            log.debug(type(worksheet))
            log.debug(f"   Worksheet name: {sheetname}")
            for validationname, validation in validation.items():
                log.debug(f"      Validation name: {validationname}")
                apply_worksheet_validation_rule(worksheet, validation)

        # Then restore the conditional formatting logic that is somehow overwritten when replacing the above details
        for sheetname, format in formatting_rules.get(template_fn.stem, {}).items():
            # Load worksheet
            worksheet = workbook[sheetname]
            log.debug(type(worksheet))
            log.debug(f"   Worksheet name: {sheetname}")
            for formattingnname, format in format.items():
                log.debug(f"      Formatting name: {formattingnname}")
                apply_worksheet_conditional_formatting(worksheet, format)

        # Define output path
        output_path = output_folder / template_fn.name

        # Save the modified workbook
        workbook.save(output_path)

    log.info(divider)
