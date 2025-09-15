import logging
import warnings
from pathlib import Path

import yaml
from openpyxl import load_workbook

from warehouse.configure.configure import get_configuration_value
from warehouse.lib.general import identify_path_by_search, produce_dir
from warehouse.lib.logging import major_header
from warehouse.lib.regex import Regex_patterns
from warehouse.lib.spreadsheets import (
    apply_worksheet_conditional_formatting,
    apply_worksheet_validation_rule,
    extract_values_from_named_range,
)
from warehouse.lib.strings import get_initials

# The exact message to ignore
warnings.filterwarnings("ignore", category=UserWarning)

# Resolve file / folder locations irrespective of cwd
script_dir = Path(__file__).parent.resolve()
templates_dir = script_dir.parent.parent.parent / "templates"


def templates(group_name: str, output_folder: Path):
    """
    Update NOMADS templates with user and project names
    """

    # Set up child log
    log = logging.getLogger(script_dir.stem)
    major_header(log, "Generating templates:")

    # Load data validation details from YAML file
    data_validations_yaml = script_dir / "data_validations.yml"
    with open(data_validations_yaml, "r") as f:
        validations = yaml.safe_load(f)

    # Load conditional formatting details from YAML file
    conditional_formatting_yaml = script_dir / "conditional_formatting.yml"
    with open(conditional_formatting_yaml, "r") as f:
        formatting_rules = yaml.safe_load(f)

    # Identify all template files
    template_fns = identify_path_by_search(
        folder_path=templates_dir, pattern=Regex_patterns.EXCEL_FILE, files_only=True
    )
    # Limit to those relevent to the group:
    template_fns = [
        t for t in template_fns if t.stem in get_configuration_value("templates")
    ]

    # Create the output folder
    produce_dir(output_folder)

    # For each template change the names, initials and projects
    for template_fn in template_fns:
        # Load the workbook
        ref_workbook = load_workbook(template_fn)
        master_version = extract_values_from_named_range(ref_workbook, "exp_version")[0]
        curr_template = output_folder / template_fn.name
        if curr_template.exists():
            curr_workbook = load_workbook(curr_template)
            curr_version = extract_values_from_named_range(
                curr_workbook, "exp_version"
            )[0]
            if master_version <= curr_version:
                log.debug(
                    f"{template_fn.stem} already exists in {output_folder} and is the latest version ({master_version})"
                )
                continue

        # Select the correct worksheet
        worksheet = ref_workbook["Reference"]
        grp_details = [
            get_configuration_value("names"),
            [get_initials(name) for name in get_configuration_value("names")],
            get_configuration_value("projects"),
        ]
        # Names, initials and projects are in cols I,J and L
        for detail_list, col_num in zip(grp_details, [9, 10, 12]):
            # Pad the list to a length of 10
            details = detail_list + [""] * (10 - len(detail_list))

            # Names, initials and projects are in rows 17 to 26
            for count, xl_row in enumerate(range(17, 27)):
                # Define the cell to edit
                cell = worksheet.cell(row=xl_row, column=col_num)
                # Assign the new value to the cell
                cell.value = details[count]

        # Then restore the data validation logic that is somehow overwritten when replacing the above details
        log.debug(f"{template_fn.stem}")

        for sheetname, validation in validations.get(template_fn.stem, {}).items():
            # Load worksheet
            worksheet = ref_workbook[sheetname]
            log.debug(type(worksheet))
            log.debug(f"   Worksheet name: {sheetname}")
            for validationname, validation in validation.items():
                log.debug(f"      Validation name: {validationname}")
                apply_worksheet_validation_rule(worksheet, validation)

        # Then restore the conditional formatting logic that is somehow overwritten when replacing the above details
        for sheetname, format in formatting_rules.get(template_fn.stem, {}).items():
            # Load worksheet
            worksheet = ref_workbook[sheetname]
            log.debug(type(worksheet))
            log.debug(f"   Worksheet name: {sheetname}")
            for formattingnname, format in format.items():
                log.debug(f"      Formatting name: {formattingnname}")
                apply_worksheet_conditional_formatting(worksheet, format)

        # Define output path
        output_path = Path(output_folder) / template_fn.name
        log.info(f"{template_fn.name} updated and output to {output_folder}")
        # Save the modified workbook
        ref_workbook.save(output_path)
